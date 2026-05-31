"""
Shopee Price Sync — Writer
==========================

Ghi kết quả ra file Excel ĐÚNG ĐỊNH DẠNG Shopee để import lại.

Nguyên tắc an toàn:
  * CHỈ ghi vào cột Giá (COL_PRICE). Không đụng cột nào khác.
  * Giữ nguyên toàn bộ header, style, sheet protection của file gốc.
  * Tô màu nhẹ những ô giá đã thay đổi để người dùng dễ rà soát (tuỳ chọn).
  * Xoá giá (ô rỗng) cho các sản phẩm ngoài batch -> Shopee hiểu là không sửa.

Đồng thời xuất:
  * report.xlsx  : các sheet Điều chỉnh / Nhóm vi phạm / Cảnh báo tồn kho.
"""

from __future__ import annotations

import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .engine import SyncResult, COL_PRICE, repair_shopee_xlsx

# màu highlight
FILL_UP = PatternFill("solid", fgColor="DCFCE7")      # xanh lá nhạt: tăng
FILL_DOWN = PatternFill("solid", fgColor="FEE2E2")    # đỏ nhạt: giảm
FILL_CLEARED = PatternFill("solid", fgColor="F1F5F9") # xám: xoá
FILL_CIRCLE = PatternFill("solid", fgColor="FEF9C3")  # vàng nhạt: ⭕


def write_output(
    original_xlsx: str | Path,
    result: SyncResult,
    out_path: str | Path,
    *,
    col_price: int = COL_PRICE,
    highlight: bool = True,
    unlock_protection: bool = True,
) -> str:
    """Ghi file import Shopee. Chỉ sửa cột giá."""
    original_xlsx = Path(original_xlsx)
    out_path = Path(out_path)

    # sửa XML lỗi trước khi mở
    tmp = out_path.with_suffix(".tmp.xlsx")
    repair_shopee_xlsx(original_xlsx, tmp)
    wb = load_workbook(tmp)
    ws = wb.active

    # Mở khoá để import không bị chặn (Shopee cho phép; người dùng có thể tắt)
    if unlock_protection and ws.protection.sheet:
        ws.protection.sheet = False

    for rec in result.rows:
        cell = ws.cell(rec.row, col_price)
        if rec.cleared:
            cell.value = None
            if highlight:
                cell.fill = FILL_CLEARED
            continue
        if rec.final_price is None:
            continue
        new_val = int(round(rec.final_price))
        old_val = None if rec.old_price is None else int(round(rec.old_price))
        cell.value = new_val
        cell.number_format = "#,##0"
        if highlight and old_val is not None and new_val != old_val:
            if rec.is_circle:
                cell.fill = FILL_CIRCLE
            elif new_val > old_val:
                cell.fill = FILL_UP
            else:
                cell.fill = FILL_DOWN
        elif highlight and rec.is_circle:
            cell.fill = FILL_CIRCLE

    wb.save(out_path)
    tmp.unlink(missing_ok=True)
    return str(out_path)


def write_report(result: SyncResult, out_path: str | Path) -> str:
    """Xuất báo cáo điều chỉnh ra file Excel nhiều sheet."""
    from openpyxl import Workbook

    out_path = Path(out_path)
    wb = Workbook()

    head_font = Font(bold=True, color="FFFFFF", name="Arial")
    head_fill = PatternFill("solid", fgColor="0F172A")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(1, c)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def autosize(ws):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 60)

    # --- Sheet 1: Tổng quan ---
    ws = wb.active
    ws.title = "Tổng quan"
    s = result.stats
    overview = [
        ["Chỉ số", "Giá trị"],
        ["Tổng dòng dữ liệu", s.get("total_rows")],
        ["Số sản phẩm", s.get("products")],
        ["Sản phẩm trong batch", s.get("products_in_batch")],
        ["Chế độ partial (xoá giá ngoài batch)", "Có" if s.get("partial_mode") else "Không"],
        ["Số SKU khớp CSV", s.get("matched")],
        ["Số dòng ⭕", s.get("circle_rows")],
        ["Số dòng điều chỉnh giá", s.get("adjustments")],
        ["Nhóm KHÔNG xử lý được", s.get("unresolved_groups")],
        ["Cảnh báo tồn kho", s.get("stock_warnings")],
        ["SKU trong CSV không khớp Excel", s.get("unmatched_csv_skus")],
        ["Giới hạn tỉ lệ max/min", s.get("params", {}).get("ratio_limit")],
        ["Tăng tối đa", f"{s.get('params',{}).get('max_up_pct',0)*100:.0f}%"],
        ["Giảm tối đa", f"{s.get('params',{}).get('max_down_pct',0)*100:.0f}%"],
    ]
    for row in overview:
        ws.append(row)
    style_header(ws, 2)
    autosize(ws)

    # --- Sheet 2: Điều chỉnh giá ---
    ws2 = wb.create_sheet("Điều chỉnh giá")
    ws2.append(["Dòng", "Mã Sản phẩm", "SKU", "Giá cũ", "Giá mục tiêu", "Giá cuối", "Chênh lệch", "Lý do"])
    for a in result.adjustments:
        diff = None
        if a.old_price is not None and a.final_price is not None:
            diff = int(a.final_price - a.old_price)
        ws2.append([
            a.row, a.product, a.sku,
            None if a.old_price is None else int(a.old_price),
            None if a.target_price is None else int(a.target_price),
            None if a.final_price is None else int(a.final_price),
            diff, a.reason,
        ])
    style_header(ws2, 8)
    for col in ("D", "E", "F", "G"):
        for cell in ws2[col][1:]:
            cell.number_format = "#,##0"
    autosize(ws2)

    # --- Sheet 3: Nhóm KHÔNG xử lý được ---
    ws3 = wb.create_sheet("Cần quyết định")
    ws3.append(["Mã Sản phẩm", "Số biến thể", "Min", "Max", "Tỉ lệ", "Ghi chú"])
    for g in result.unresolved:
        ws3.append([
            g.product, g.n_variations,
            None if g.min_price is None else int(g.min_price),
            None if g.max_price is None else int(g.max_price),
            None if g.ratio is None else round(g.ratio, 2),
            g.note,
        ])
    style_header(ws3, 6)
    for col in ("C", "D"):
        for cell in ws3[col][1:]:
            cell.number_format = "#,##0"
    autosize(ws3)

    # --- Sheet 4: Cảnh báo tồn kho ---
    ws4 = wb.create_sheet("Cảnh báo tồn kho")
    ws4.append(["Dòng", "Mã Sản phẩm", "SKU", "Số lượng", "Loại", "Chi tiết"])
    for w in result.stock_warnings:
        ws4.append([w.row, w.product, w.sku, w.stock, w.kind, w.detail])
    style_header(ws4, 6)
    autosize(ws4)

    # --- Sheet 5: SKU CSV không khớp ---
    ws5 = wb.create_sheet("SKU CSV không khớp")
    ws5.append(["MÃ NỘI BỘ (không thấy trong Excel)"])
    for k in result.unmatched_skus:
        ws5.append([k])
    style_header(ws5, 1)
    autosize(ws5)

    wb.save(out_path)
    return str(out_path)

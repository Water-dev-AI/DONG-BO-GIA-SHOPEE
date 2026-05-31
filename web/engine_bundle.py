"""
Shopee Price Sync — Core Engine
================================

Đồng bộ giá từ bảng giá mới (CSV / Google Sheet) vào file Excel chuẩn Shopee,
đồng thời tự động xử lý ràng buộc giá max/min không vượt quá 5 lần trong cùng
một Mã Sản phẩm.

Nguyên tắc (theo yêu cầu người dùng):
  * Trong 1 Mã Sản phẩm (product group): max / min KHÔNG được vượt quá RATIO_LIMIT (5).
  * Nếu vi phạm:
      - Giá thấp hơn ngưỡng min cho phép  -> TĂNG lên (tối đa +MAX_UP_PCT, mặc định 50%).
      - Giá cao hơn ngưỡng max cho phép    -> GIẢM xuống (tối đa -MAX_DOWN_PCT, mặc định 2%).
      - Có thể vừa hạ max vừa nâng min để "gặp nhau" cho hài hoà.
  * Làm tròn tới hàng NGHÌN (10000 ok, 10500 -> 10000 / 11000 tuỳ chiều).
  * Nếu KHÔNG thể đưa về trong giới hạn -> báo cáo để người dùng tự quyết.

Quy ước cột trong Excel Shopee (data bắt đầu từ DATA_START_ROW):
  A (1) Mã Sản phẩm     -> nhóm sản phẩm để tính min/max
  C (3) Mã Phân loại
  F (6) SKU             -> = "MÃ NỘI BỘ" trong CSV (khoá vlookup)
  G (7) Giá             -> ô DUY NHẤT được sửa
  I (9) Số lượng

CSV: chỉ cần 2 cột  "MÃ NỘI BỘ"  và  "GIÁ PEE".
"""

from __future__ import annotations

import csv
import io
import math
import re
import shutil
import tempfile
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional

# ----------------------------------------------------------------------------
# Cấu hình mặc định
# ----------------------------------------------------------------------------

RATIO_LIMIT = 5.0          # max/min không quá 5 lần
ROUND_STEP = 1000          # làm tròn hàng nghìn
DEFAULT_MAX_UP_PCT = 0.50  # tăng tối đa 50% giá bán
DEFAULT_MAX_DOWN_PCT = 0.02  # giảm tối đa 2% giá bán
PRICE_FLOOR = 1000         # Shopee: giá tối thiểu
PRICE_CEIL = 120_000_000   # Shopee: giá tối đa

# Map cột Excel (1-based) — có thể override khi gọi
COL_PRODUCT = 1   # A Mã Sản phẩm
COL_VARIATION = 3 # C Mã Phân loại
COL_SKU = 6       # F SKU
COL_PRICE = 7     # G Giá  (ô duy nhất được sửa)
COL_STOCK = 9     # I Số lượng
DATA_START_ROW = 7

CIRCLE = "⭕"

CSV_SKU_HEADER_CANDIDATES = ["MÃ NỘI BỘ", "MA NOI BO", "SKU"]
CSV_PRICE_HEADER_CANDIDATES = ["GIÁ PEE", "GIA PEE", "GIÁ\nPEE", "GIÁ \nPEE", "GIÁ"]


# ----------------------------------------------------------------------------
# Sửa lỗi XML của file Shopee (activePane="bottom_left" -> "bottomLeft")
# openpyxl không đọc được file Shopee export vì lỗi này.
# ----------------------------------------------------------------------------

_PANE_FIXES = {
    'activePane="bottom_left"': 'activePane="bottomLeft"',
    'activePane="bottom_right"': 'activePane="bottomRight"',
    'activePane="top_left"': 'activePane="topLeft"',
    'activePane="top_right"': 'activePane="topRight"',
}


def repair_shopee_xlsx(src: str | Path, dst: str | Path) -> bool:
    """Sửa các giá trị enum XML không hợp lệ trong file Shopee.
    Trả về True nếu có sửa đổi."""
    src, dst = Path(src), Path(dst)
    changed = False
    with zipfile.ZipFile(src, "r") as zin:
        names = zin.namelist()
        buffers = {n: zin.read(n) for n in names}
    for n, data in buffers.items():
        if n.startswith("xl/worksheets/") and n.endswith(".xml"):
            text = data.decode("utf-8", errors="ignore")
            new = text
            for bad, good in _PANE_FIXES.items():
                if bad in new:
                    new = new.replace(bad, good)
                    changed = True
            if new != text:
                buffers[n] = new.encode("utf-8")
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for n, data in buffers.items():
            zout.writestr(n, data)
    return changed


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s)).strip().upper()


def round_thousand(x: float, mode: str = "nearest") -> int:
    """Làm tròn về hàng nghìn.
    mode: 'nearest' | 'up' | 'down'."""
    if mode == "up":
        return int(math.ceil(x / ROUND_STEP) * ROUND_STEP)
    if mode == "down":
        return int(math.floor(x / ROUND_STEP) * ROUND_STEP)
    return int(round(x / ROUND_STEP) * ROUND_STEP)


def parse_price(v) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    # bỏ ký tự tiền tệ / khoảng trắng
    s = re.sub(r"[^\d.,]", "", s)
    if s == "":
        return None
    # Định dạng VN: dấu '.' hoặc ',' thường là phân tách hàng nghìn.
    # Nếu phần sau dấu cuối cùng có đúng 3 chữ số -> đó là nghìn, bỏ hết separator.
    last_sep = max(s.rfind("."), s.rfind(","))
    if last_sep != -1:
        tail = s[last_sep + 1:]
        if len(tail) == 3 and tail.isdigit():
            s = re.sub(r"[.,]", "", s)            # tất cả là phân tách nghìn
        else:
            s = s.replace(",", "")                 # ',' nghìn, '.' thập phân
            # nếu có nhiều dấu '.', chỉ giữ dấu cuối làm thập phân
            if s.count(".") > 1:
                parts = s.split(".")
                s = "".join(parts[:-1]) + "." + parts[-1]
    try:
        return float(s)
    except ValueError:
        return None


def is_circle(sku) -> bool:
    return sku is not None and CIRCLE in str(sku)


# ----------------------------------------------------------------------------
# Đọc bảng giá mới (CSV / TSV) -> dict[sku_norm] = price
# ----------------------------------------------------------------------------

def load_price_map_from_csv_bytes(raw: bytes) -> dict:
    text = raw.decode("utf-8-sig", errors="replace")
    # auto-detect delimiter
    sample = text[:5000]
    delim = "\t" if sample.count("\t") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = list(reader)
    if not rows:
        return {}
    header = rows[0]
    norm_header = [_norm(h) for h in header]

    def find_col(cands):
        for c in cands:
            cn = _norm(c)
            for i, h in enumerate(norm_header):
                if h == cn:
                    return i
        # fuzzy contains
        for c in cands:
            cn = _norm(c)
            for i, h in enumerate(norm_header):
                if cn in h or h in cn:
                    return i
        return None

    sku_i = find_col(CSV_SKU_HEADER_CANDIDATES)
    price_i = find_col(CSV_PRICE_HEADER_CANDIDATES)
    if sku_i is None or price_i is None:
        raise ValueError(
            f"Không tìm thấy cột MÃ NỘI BỘ / GIÁ PEE trong CSV. Header: {header}"
        )
    out = {}
    for r in rows[1:]:
        if len(r) <= max(sku_i, price_i):
            continue
        sku = str(r[sku_i]).strip()
        if not sku:
            continue
        price = parse_price(r[price_i])
        if price is None:
            continue
        out[_norm(sku)] = price
    return out


def load_price_map_from_path(path: str | Path) -> dict:
    return load_price_map_from_csv_bytes(Path(path).read_bytes())


# ----------------------------------------------------------------------------
# Data classes cho kết quả
# ----------------------------------------------------------------------------

@dataclass
class RowRecord:
    row: int
    product: str
    variation: str
    sku: str
    is_circle: bool
    old_price: Optional[float]
    new_price: Optional[float]      # giá target trước constraint
    final_price: Optional[float]    # giá sau khi giải constraint + làm tròn
    stock: Optional[float]
    matched_csv: bool = False
    cleared: bool = False           # bị xoá giá (sản phẩm không nằm trong batch)


@dataclass
class GroupReport:
    product: str
    n_variations: int
    min_price: Optional[float]
    max_price: Optional[float]
    ratio: Optional[float]
    resolved: bool
    note: str = ""


@dataclass
class Adjustment:
    row: int
    product: str
    sku: str
    old_price: Optional[float]
    target_price: Optional[float]
    final_price: Optional[float]
    reason: str


@dataclass
class StockWarning:
    row: int
    product: str
    sku: str
    stock: Optional[float]
    kind: str  # 'circle_nonzero' | 'noncircle_zero'
    detail: str


@dataclass
class SyncResult:
    rows: list = field(default_factory=list)
    groups: list = field(default_factory=list)
    adjustments: list = field(default_factory=list)
    stock_warnings: list = field(default_factory=list)
    unresolved: list = field(default_factory=list)   # GroupReport không giải được
    unmatched_skus: list = field(default_factory=list)
    stats: dict = field(default_factory=dict)

    def to_dict(self):
        return {
            "rows": [asdict(r) for r in self.rows],
            "groups": [asdict(g) for g in self.groups],
            "adjustments": [asdict(a) for a in self.adjustments],
            "stock_warnings": [asdict(s) for s in self.stock_warnings],
            "unresolved": [asdict(u) for u in self.unresolved],
            "unmatched_skus": self.unmatched_skus,
            "stats": self.stats,
        }


# ----------------------------------------------------------------------------
# Bộ giải ràng buộc max/min cho 1 nhóm
# ----------------------------------------------------------------------------

def solve_group(
    prices: list[float],
    max_up_pct: float,
    max_down_pct: float,
    ratio_limit: float = RATIO_LIMIT,
) -> tuple[list[float], bool, str]:
    """
    Nhận list giá target (đã làm tròn nghìn) của 1 nhóm.
    Trả về (final_prices, resolved, note).

    Chiến lược:
      Nếu max/min <= ratio_limit  -> giữ nguyên.
      Ngược lại cần thu hẹp khoảng [min, max] sao cho max' <= min' * ratio_limit.
      Ràng buộc dịch chuyển mỗi item:
          lower_i = round_up(price_i * (1 - max_down_pct))   # giảm tối đa 2%
          upper_i = round_down(price_i * (1 + max_up_pct))   # tăng tối đa 50%
      Ta tìm một dải [L, U] với U <= L*ratio_limit sao cho mỗi giá có thể
      được kéo vào trong [L, U] (clamp) trong giới hạn cho phép.

      Cách làm "hài hoà": kéo các giá CAO xuống tối đa cho phép, kéo các giá
      THẤP lên tối đa cho phép, rồi kiểm tra ratio. Nếu vẫn vi phạm thì
      không giải được -> báo cáo.
    """
    if not prices:
        return [], True, ""
    n = len(prices)
    cur = list(prices)
    mn, mx = min(cur), max(cur)
    if mn <= 0:
        return cur, False, "Có giá <= 0"
    if mx <= mn * ratio_limit + 1e-9:
        return cur, True, ""  # đã hợp lệ

    # biên cho phép của mỗi item
    def lo_bound(p):  # giảm tối đa max_down_pct
        return round_thousand(p * (1 - max_down_pct), "up")  # giảm ít nhất -> làm tròn lên

    def hi_bound(p):  # tăng tối đa max_up_pct
        return round_thousand(p * (1 + max_up_pct), "down")

    los = [lo_bound(p) for p in cur]  # giá thấp nhất có thể hạ xuống
    his = [hi_bound(p) for p in cur]  # giá cao nhất có thể nâng lên

    # Khoảng khả thi tổng: giá nhỏ nhất có thể của max-item = its lo
    # giá lớn nhất có thể của min-item = its hi
    # Cần tồn tại L, U: L <= U <= L*ratio, và với mỗi i: clamp(cur_i, L, U) khả thi
    #   -> nếu cur_i > U thì phải hạ về U, cần U >= los[i]
    #   -> nếu cur_i < L thì phải nâng về L, cần L <= his[i]
    #
    # Đặt L = giá nhỏ nhất ta chấp nhận, U = L*ratio.
    # Thử L chạy từ max(các giá phải nâng) — dùng tìm kiếm trên tập ứng viên.

    candidates = sorted(set(
        [round_thousand(p, "nearest") for p in cur]
        + los + his
        + [round_thousand(mn, "nearest"), round_thousand(mx / ratio_limit, "up")]
    ))

    best = None
    for L in candidates:
        if L < PRICE_FLOOR:
            continue
        U = math.floor((L * ratio_limit) / ROUND_STEP) * ROUND_STEP
        if U < L:
            continue
        feasible = True
        proposal = []
        cost = 0.0
        for i, p in enumerate(cur):
            if p < L:               # cần nâng lên L
                if L > his[i]:      # vượt trần tăng 50%
                    feasible = False
                    break
                newp = L
            elif p > U:             # cần hạ xuống U
                if U < los[i]:      # vượt sàn giảm 2%
                    feasible = False
                    break
                newp = U
            else:
                newp = p
            proposal.append(newp)
            cost += abs(newp - p)
        if feasible:
            if best is None or cost < best[1]:
                best = (proposal, cost, L, U)
    if best is None:
        return cur, False, (
            f"Không thể đưa max/min về <= {ratio_limit:g} lần "
            f"trong giới hạn (tăng {max_up_pct*100:.0f}% / giảm {max_down_pct*100:.0f}%)."
        )
    proposal = best[0]
    # đảm bảo làm tròn nghìn
    proposal = [round_thousand(x, "nearest") for x in proposal]
    # double-check ratio sau làm tròn
    m2, x2 = min(proposal), max(proposal)
    note = ""
    if x2 > m2 * ratio_limit + 1e-9:
        return proposal, False, "Sau làm tròn vẫn vượt giới hạn — cần xem lại thủ công."
    return proposal, True, note


# ----------------------------------------------------------------------------
# Engine chính
# ----------------------------------------------------------------------------

def run_sync(
    xlsx_path: str | Path,
    price_map: dict,
    *,
    max_up_pct: float = DEFAULT_MAX_UP_PCT,
    max_down_pct: float = DEFAULT_MAX_DOWN_PCT,
    ratio_limit: float = RATIO_LIMIT,
    clear_unmatched: bool = True,
    col_product: int = COL_PRODUCT,
    col_sku: int = COL_SKU,
    col_price: int = COL_PRICE,
    col_stock: int = COL_STOCK,
    col_variation: int = COL_VARIATION,
    data_start_row: int = DATA_START_ROW,
) -> tuple[SyncResult, str]:
    """
    Đọc Excel Shopee, áp giá mới, giải ràng buộc, trả về (SyncResult, fixed_xlsx_path).
    KHÔNG ghi file output ở đây — chỉ tính toán. (writer.py lo phần ghi.)
    """
    from openpyxl import load_workbook

    xlsx_path = Path(xlsx_path)
    tmp_fixed = Path(tempfile.mkdtemp()) / "fixed.xlsx"
    repair_shopee_xlsx(xlsx_path, tmp_fixed)
    wb = load_workbook(tmp_fixed)
    ws = wb.active

    result = SyncResult()
    rows: list[RowRecord] = []

    # batch = tập product có ít nhất 1 SKU khớp CSV
    products_in_batch = set()

    # 1) Đọc toàn bộ + áp target price
    for r in range(data_start_row, ws.max_row + 1):
        pid = ws.cell(r, col_product).value
        sku = ws.cell(r, col_sku).value
        price = parse_price(ws.cell(r, col_price).value)
        stock = ws.cell(r, col_stock).value
        var = ws.cell(r, col_variation).value
        if pid is None and sku is None and price is None and stock is None:
            continue
        pid_s = "" if pid is None else str(pid).strip()
        sku_s = "" if sku is None else str(sku).strip()
        circ = is_circle(sku_s)
        try:
            stock_f = float(stock) if stock not in (None, "") else None
        except (ValueError, TypeError):
            stock_f = None

        matched = False
        target = price
        if not circ and sku_s:
            key = _norm(sku_s)
            if key in price_map:
                target = price_map[key]
                matched = True
                products_in_batch.add(pid_s)

        rows.append(RowRecord(
            row=r, product=pid_s, variation="" if var is None else str(var),
            sku=sku_s, is_circle=circ, old_price=price,
            new_price=target, final_price=target, stock=stock_f,
            matched_csv=matched,
        ))

    # 2) Xác định chế độ: full-list hay partial-batch
    #    Nếu clear_unmatched: các product KHÔNG nằm trong batch -> xoá giá (Shopee bỏ qua).
    #    Product trong batch nhưng có SKU không khớp CSV -> giữ giá cũ làm target.
    partial_mode = clear_unmatched and len(products_in_batch) < _count_products(rows)

    # 3) Gom nhóm và giải ràng buộc (chỉ với product nằm trong batch / full)
    groups = defaultdict(list)
    for rec in rows:
        groups[rec.product].append(rec)

    for pid, recs in groups.items():
        in_batch = pid in products_in_batch
        if partial_mode and not in_batch:
            # product không điều chỉnh -> sẽ xoá giá ở writer
            for rec in recs:
                rec.cleared = True
                rec.final_price = None
            result.groups.append(GroupReport(
                product=pid, n_variations=len(recs),
                min_price=None, max_price=None, ratio=None,
                resolved=True, note="Không nằm trong batch — xoá giá để Shopee bỏ qua.",
            ))
            continue

        # các SKU non-circle có target để giải ràng buộc
        priced = [rec for rec in recs if not rec.is_circle and rec.new_price]
        # làm tròn nghìn cho mọi target trước
        for rec in priced:
            rec.new_price = float(round_thousand(rec.new_price, "nearest"))
            rec.final_price = rec.new_price

        if not priced:
            result.groups.append(GroupReport(
                product=pid, n_variations=len(recs),
                min_price=None, max_price=None, ratio=None,
                resolved=True, note="Không có giá hợp lệ.",
            ))
            continue

        target_prices = [rec.new_price for rec in priced]
        finals, resolved, note = solve_group(
            target_prices, max_up_pct, max_down_pct, ratio_limit
        )

        for rec, fp in zip(priced, finals):
            rec.final_price = float(fp)
            if rec.old_price is None or abs((rec.final_price or 0) - (rec.old_price or 0)) > 0.5 \
               or (rec.matched_csv and rec.new_price != rec.old_price):
                pass  # adjustments tổng hợp ở dưới

        gmin, gmax = min(finals), max(finals)
        gratio = gmax / gmin if gmin else None
        report = GroupReport(
            product=pid, n_variations=len(recs),
            min_price=gmin, max_price=gmax, ratio=gratio,
            resolved=resolved, note=note,
        )
        result.groups.append(report)
        if not resolved:
            result.unresolved.append(report)

        # 4) Gán giá cho ⭕ = max của nhóm (sau khi giải)
        group_max = gmax
        for rec in recs:
            if rec.is_circle:
                rec.final_price = float(group_max)

    # 5) Tổng hợp adjustments (chỉ những dòng thực sự đổi giá)
    for rec in rows:
        if rec.cleared:
            result.adjustments.append(Adjustment(
                row=rec.row, product=rec.product, sku=rec.sku,
                old_price=rec.old_price, target_price=None, final_price=None,
                reason="Xoá giá (sản phẩm ngoài batch)",
            ))
            continue
        if rec.old_price is None and rec.final_price is None:
            continue
        op = rec.old_price
        fp = rec.final_price
        if op is None or fp is None or abs(fp - op) > 0.5:
            reason = _adjust_reason(rec)
            result.adjustments.append(Adjustment(
                row=rec.row, product=rec.product, sku=rec.sku,
                old_price=op, target_price=rec.new_price, final_price=fp,
                reason=reason,
            ))

    # 6) Kiểm tra tồn kho / ⭕
    for rec in rows:
        if rec.is_circle:
            if rec.stock not in (0, 0.0, None):
                result.stock_warnings.append(StockWarning(
                    row=rec.row, product=rec.product, sku=rec.sku, stock=rec.stock,
                    kind="circle_nonzero",
                    detail="SKU ⭕ nhưng số lượng khác 0 (cần kiểm tra).",
                ))
        else:
            if rec.stock in (0, 0.0):
                result.stock_warnings.append(StockWarning(
                    row=rec.row, product=rec.product, sku=rec.sku, stock=rec.stock,
                    kind="noncircle_zero",
                    detail="SKU thường nhưng số lượng = 0 (có thể sai số lượng).",
                ))

    # 7) SKU trong CSV không khớp Excel
    excel_skus = {_norm(rec.sku) for rec in rows if rec.sku and not rec.is_circle}
    for k in price_map:
        if k not in excel_skus:
            result.unmatched_skus.append(k)

    result.rows = rows
    result.stats = {
        "total_rows": len(rows),
        "products": _count_products(rows),
        "products_in_batch": len(products_in_batch),
        "partial_mode": partial_mode,
        "matched": sum(1 for r in rows if r.matched_csv),
        "circle_rows": sum(1 for r in rows if r.is_circle),
        "adjustments": len(result.adjustments),
        "unresolved_groups": len(result.unresolved),
        "stock_warnings": len(result.stock_warnings),
        "unmatched_csv_skus": len(result.unmatched_skus),
        "params": {
            "ratio_limit": ratio_limit,
            "max_up_pct": max_up_pct,
            "max_down_pct": max_down_pct,
            "round_step": ROUND_STEP,
            "clear_unmatched": clear_unmatched,
        },
    }
    return result, str(tmp_fixed)


def _count_products(rows: list[RowRecord]) -> int:
    return len({r.product for r in rows if r.product})


def _adjust_reason(rec: RowRecord) -> str:
    op, fp, tp = rec.old_price, rec.final_price, rec.new_price
    parts = []
    if rec.is_circle:
        return "⭕ = max nhóm"
    if rec.matched_csv:
        parts.append("Cập nhật giá mới")
    if tp is not None and fp is not None and abs(fp - tp) > 0.5:
        if fp > tp:
            parts.append("nâng min (ràng buộc 5x)")
        else:
            parts.append("hạ max (ràng buộc 5x)")
    elif op is not None and fp is not None and abs(fp - op) > 0.5 and not rec.matched_csv:
        parts.append("điều chỉnh ràng buộc 5x")
    return " + ".join(parts) if parts else "Cập nhật"


# ====== WRITER ======
"""
Shopee Price Sync — Writer
==========================

Ghi kết quả ra file Excel ĐÚNG ĐỊNH DẠNG Shopee để import lại.

Nguyên tắc an toàn:
  * CHỈ ghi vào cột Giá (COL_PRICE). Không đụng cột nào khác.
  * Giữ nguyên toàn bộ header, style, sheet protection của file gốc.
  * Tô màu nhẹ những ô giá đã thay đổi để người dùng dễ rà soát (tuỳ chọn).
  * Xoá giá (ô rỗng) cho các sản phẩm ngoài batch -> Shopee hiểu là không sửa.

Đồng thời xuất:
  * report.xlsx  : các sheet Điều chỉnh / Nhóm vi phạm / Cảnh báo tồn kho.
"""



from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side



# màu highlight
FILL_UP = PatternFill("solid", fgColor="DCFCE7")      # xanh lá nhạt: tăng
FILL_DOWN = PatternFill("solid", fgColor="FEE2E2")    # đỏ nhạt: giảm
FILL_CLEARED = PatternFill("solid", fgColor="F1F5F9") # xám: xoá
FILL_CIRCLE = PatternFill("solid", fgColor="FEF9C3")  # vàng nhạt: ⭕


def write_output(
    original_xlsx: str | Path,
    result: SyncResult,
    out_path: str | Path,
    *,
    col_price: int = COL_PRICE,
    highlight: bool = True,
    unlock_protection: bool = True,
) -> str:
    """Ghi file import Shopee. Chỉ sửa cột giá."""
    original_xlsx = Path(original_xlsx)
    out_path = Path(out_path)

    # sửa XML lỗi trước khi mở
    tmp = out_path.with_suffix(".tmp.xlsx")
    repair_shopee_xlsx(original_xlsx, tmp)
    wb = load_workbook(tmp)
    ws = wb.active

    # Mở khoá để import không bị chặn (Shopee cho phép; người dùng có thể tắt)
    if unlock_protection and ws.protection.sheet:
        ws.protection.sheet = False

    for rec in result.rows:
        cell = ws.cell(rec.row, col_price)
        if rec.cleared:
            cell.value = None
            if highlight:
                cell.fill = FILL_CLEARED
            continue
        if rec.final_price is None:
            continue
        new_val = int(round(rec.final_price))
        old_val = None if rec.old_price is None else int(round(rec.old_price))
        cell.value = new_val
        cell.number_format = "#,##0"
        if highlight and old_val is not None and new_val != old_val:
            if rec.is_circle:
                cell.fill = FILL_CIRCLE
            elif new_val > old_val:
                cell.fill = FILL_UP
            else:
                cell.fill = FILL_DOWN
        elif highlight and rec.is_circle:
            cell.fill = FILL_CIRCLE

    wb.save(out_path)
    tmp.unlink(missing_ok=True)
    return str(out_path)


def write_report(result: SyncResult, out_path: str | Path) -> str:
    """Xuất báo cáo điều chỉnh ra file Excel nhiều sheet."""
    from openpyxl import Workbook

    out_path = Path(out_path)
    wb = Workbook()

    head_font = Font(bold=True, color="FFFFFF", name="Arial")
    head_fill = PatternFill("solid", fgColor="0F172A")
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(1, c)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def autosize(ws):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 60)

    # --- Sheet 1: Tổng quan ---
    ws = wb.active
    ws.title = "Tổng quan"
    s = result.stats
    overview = [
        ["Chỉ số", "Giá trị"],
        ["Tổng dòng dữ liệu", s.get("total_rows")],
        ["Số sản phẩm", s.get("products")],
        ["Sản phẩm trong batch", s.get("products_in_batch")],
        ["Chế độ partial (xoá giá ngoài batch)", "Có" if s.get("partial_mode") else "Không"],
        ["Số SKU khớp CSV", s.get("matched")],
        ["Số dòng ⭕", s.get("circle_rows")],
        ["Số dòng điều chỉnh giá", s.get("adjustments")],
        ["Nhóm KHÔNG xử lý được", s.get("unresolved_groups")],
        ["Cảnh báo tồn kho", s.get("stock_warnings")],
        ["SKU trong CSV không khớp Excel", s.get("unmatched_csv_skus")],
        ["Giới hạn tỉ lệ max/min", s.get("params", {}).get("ratio_limit")],
        ["Tăng tối đa", f"{s.get('params',{}).get('max_up_pct',0)*100:.0f}%"],
        ["Giảm tối đa", f"{s.get('params',{}).get('max_down_pct',0)*100:.0f}%"],
    ]
    for row in overview:
        ws.append(row)
    style_header(ws, 2)
    autosize(ws)

    # --- Sheet 2: Điều chỉnh giá ---
    ws2 = wb.create_sheet("Điều chỉnh giá")
    ws2.append(["Dòng", "Mã Sản phẩm", "SKU", "Giá cũ", "Giá mục tiêu", "Giá cuối", "Chênh lệch", "Lý do"])
    for a in result.adjustments:
        diff = None
        if a.old_price is not None and a.final_price is not None:
            diff = int(a.final_price - a.old_price)
        ws2.append([
            a.row, a.product, a.sku,
            None if a.old_price is None else int(a.old_price),
            None if a.target_price is None else int(a.target_price),
            None if a.final_price is None else int(a.final_price),
            diff, a.reason,
        ])
    style_header(ws2, 8)
    for col in ("D", "E", "F", "G"):
        for cell in ws2[col][1:]:
            cell.number_format = "#,##0"
    autosize(ws2)

    # --- Sheet 3: Nhóm KHÔNG xử lý được ---
    ws3 = wb.create_sheet("Cần quyết định")
    ws3.append(["Mã Sản phẩm", "Số biến thể", "Min", "Max", "Tỉ lệ", "Ghi chú"])
    for g in result.unresolved:
        ws3.append([
            g.product, g.n_variations,
            None if g.min_price is None else int(g.min_price),
            None if g.max_price is None else int(g.max_price),
            None if g.ratio is None else round(g.ratio, 2),
            g.note,
        ])
    style_header(ws3, 6)
    for col in ("C", "D"):
        for cell in ws3[col][1:]:
            cell.number_format = "#,##0"
    autosize(ws3)

    # --- Sheet 4: Cảnh báo tồn kho ---
    ws4 = wb.create_sheet("Cảnh báo tồn kho")
    ws4.append(["Dòng", "Mã Sản phẩm", "SKU", "Số lượng", "Loại", "Chi tiết"])
    for w in result.stock_warnings:
        ws4.append([w.row, w.product, w.sku, w.stock, w.kind, w.detail])
    style_header(ws4, 6)
    autosize(ws4)

    # --- Sheet 5: SKU CSV không khớp ---
    ws5 = wb.create_sheet("SKU CSV không khớp")
    ws5.append(["MÃ NỘI BỘ (không thấy trong Excel)"])
    for k in result.unmatched_skus:
        ws5.append([k])
    style_header(ws5, 1)
    autosize(ws5)

    wb.save(out_path)
    return str(out_path)


# ============================================================================
# Glue cho Pyodide — process_all() được gọi từ JavaScript
# ============================================================================
import base64 as _b64
import tempfile as _tf
from pathlib import Path as _P


def process_all():
    """Đọc XLSX_BYTES / CSV_BYTES (globals từ JS), trả dict kết quả + file b64."""
    xb = bytes(XLSX_BYTES)
    cb = bytes(CSV_BYTES)
    up = float(P_UP); down = float(P_DOWN); ratio = float(P_RATIO)
    rstep = int(P_ROUND); clear = bool(P_CLEAR); unlock = bool(P_UNLOCK)

    global ROUND_STEP
    ROUND_STEP = rstep

    d = _P(_tf.mkdtemp())
    xin = d / "in.xlsx"
    xin.write_bytes(xb)

    pm = load_price_map_from_csv_bytes(cb)
    result, _ = run_sync(
        str(xin), pm,
        max_up_pct=up, max_down_pct=down, ratio_limit=ratio,
        clear_unmatched=clear,
    )

    out = d / "import_shopee.xlsx"
    write_output(str(xin), result, str(out), unlock_protection=unlock)
    rep = d / "report.xlsx"
    write_report(result, str(rep))

    payload = result.to_dict()
    payload["import_b64"] = _b64.b64encode(out.read_bytes()).decode()
    payload["report_b64"] = _b64.b64encode(rep.read_bytes()).decode()
    # rows quá lớn, không cần gửi sang JS
    payload.pop("rows", None)
    payload.pop("groups", None)
    return payload
